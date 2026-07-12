import os
import requests
import time
import math
import yfinance as yf
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# ============================================================
# EMBEDDED LIST 1: CRYPTOCURRENCY CONFIG
# ============================================================
COIN_LIST = [
    "bitcoin", "ethereum", "solana", "ripple", "cardano",
    "dogecoin", "shiba-inu", "pepe", "paxg"
]

# ============================================================
# EMBEDDED LIST 2: STOCK CONFIG
# ============================================================
STOCK_LIST = [
    "NVDA", "AMD", "INTC", "MSTR", "ORCL", "AI", "GTLB", "TEM", "CGNX",
    "SNOW", "DDOG", "IPGP", "MU", "ASX", "AMKR", "ANET", "WDC", "AVGO",
    "MRVL", "KEYS", "SNDK", "CSCO", "QCOM", "ARM"
]

# ============================================================
# COMBINED ATH DATA
# ============================================================
ATH_DATA = {
    "BITCOIN":   {"ath": 126198.10,  "date": "10/05/2025"},
    "ETHEREUM":  {"ath": 4953.73,    "date": "08/24/2025"},
    "SOLANA":    {"ath": 294.33,     "date": "01/19/2025"},
    "RIPPLE":    {"ath": 3.84,       "date": "01/13/2018"},
    "CARDANO":   {"ath": 3.10,       "date": "09/02/2021"},
    "DOGECOIN":  {"ath": 0.7375,     "date": "05/08/2021"},
    "SHIBA-INU": {"ath": 0.00008845, "date": "10/28/2021"},
    "PEPE":      {"ath": 0.00002824, "date": "12/09/2024"},
    "PAXG":      {"ath": 2855.83,    "date": "04/13/2024"},
    "NVDA":  {"ath": 235.74,  "date": "05/14/2026"},
    "AMD":   {"ath": 584.73,  "date": "06/30/2026"},
    "INTC":  {"ath": 142.35,  "date": "06/30/2026"},
    "MSTR":  {"ath": 457.22,  "date": "1/1/2025"},
    "ORCL":  {"ath": 345.72,  "date": "09/10/2025"},
    "AI":    {"ath": 30.11,  "date": "1/1/2025"},
    "GTLB":  {"ath": 52.38,  "date": "11/03/2025"},
    "TEM":   {"ath": 104.32,  "date": "11/09/2025"},
    "CGNX":  {"ath": 101.82,  "date": "02/12/2021"},
    "SNOW":  {"ath": 280.16,  "date": "06/01/2026"},
    "DDOG":  {"ath": 278.70,  "date": "06/01/2026"},
    "IPGP":  {"ath": 155.82,  "date": "1/1/2026"},
    "MU":    {"ath": 1213.56,  "date": "06/25/2026"},
    "ASX":   {"ath": 45.32,  "date": "06/30/2026"},
    "AMKR":  {"ath": 96.29,  "date": "06/22/2026"},
    "ANET":  {"ath": 179.80,  "date": "1/1/2026"},
    "WDC":   {"ath": 799.87,  "date": "06/18/2026"},
    "AVGO":  {"ath": 495.00,  "date": "06/03/2026"},
    "MRVL":  {"ath": 329.88,  "date": "06/18/2026"},
    "KEYS":  {"ath": 373.34,  "date": "06/22/2026"},
    "SNDK":  {"ath": 2348.00,  "date": "06/25/2026"},
    "CSCO":  {"ath": 130.37,  "date": "06/04/2026"},
    "QCOM":  {"ath": 259.92,  "date": "05/29/2026"},
    "ARM":  {"ath": 439.46,  "date": "0618/2026"},
}

NUM_WEEKS  = 52
NUM_MONTHS = 12
NUM_DAYS   = 30
EXCEL_FILE = "market_data_greport.xlsx"
SHEET_ROW  = "Row Report"
SHEET_COL  = "Column Report"

# ============================================================
# UTILITY FUNCTIONS
# ============================================================

def format_price(price):
    if isinstance(price, str): return price
    if price < 0.01: return f"${price:.8f}"
    return f"${price:.2f}"

def format_excel_num(val):
    if val == "" or val is None:  return ""
    if isinstance(val, str):      return val
    if isinstance(val, float) and val < 0.01: return f"{val:.8f}"
    return round(val, 4)

def safe_ratio(min_val, max_val):
    try:
        if min_val is None or max_val is None: return ""
        if min_val == "" or max_val == "":     return ""
        if max_val == 0:                       return ""
        return round(min_val / max_val, 6)
    except Exception:
        return ""

def are_equal_prices(a, b):
    if not isinstance(a, (int, float)): return False
    if not isinstance(b, (int, float)): return False
    if a == 0 and b == 0: return True
    tolerance = max(abs(a), abs(b)) * 1e-6
    return abs(a - b) <= tolerance

def parse_price_for_match(val):
    """Safely converts stringified floats back to float for matching."""
    if isinstance(val, (int, float)): return val
    if isinstance(val, str):
        try:
            return float(val.replace('$', '').replace(',', ''))
        except Exception:
            return None
    return None

def calculate_metrics(min_val, max_val):
    if isinstance(min_val, str) or isinstance(max_val, str) or max_val == 0:
        return "ERROR", "ERROR", "ERROR", "ERROR"
    ratio       = min_val / max_val
    ref         = math.floor((1.0 - ratio) * 100) / 100.0
    target_down = max_val * (1 - 2 * ref)
    target_up   = min_val * (1 + 2 * ref)
    return ratio, ref, target_down, target_up

def get_tracking_dates():
    """Generates date headers starting from 06/22/2026 up to today, in reverse chronological order."""
    start_date = datetime(2026, 6, 22)
    now        = datetime.now()
    dates      = []
    current    = start_date
    while current.date() <= now.date():
        dates.append(f"{current.month}/{current.day}/{current.year}")
        current += timedelta(days=1)
        
    # Reverse the list so the newest date is first (closest to the daily columns)
    dates.reverse()
    return dates

def get_color_rgb(color_obj):
    """Safely extracts strict 8-character aRGB hex strings, ignoring invalid/theme colors."""
    if not color_obj: return None
    try:
        if getattr(color_obj, 'type', None) == 'rgb':
            val = str(color_obj.rgb)
            if len(val) == 8 and val != "00000000":
                return val
    except:
        pass
    return None

# ============================================================
# EXCEL HEADER BUILDER  (shared by both sheets)
# ============================================================

def build_excel_headers():
    base = [
        "ASSET", "CURRENT",
        "7D MIN", "7D MAX", "7D min/Max Ratio",
        "30D MIN", "30D MAX", "30D min/Max Ratio",
        "1Y MIN", "1Y MAX", "1yr min/Max Ratio",
        "1yr ref", "1yr target down", "1yr target up",
        "% to ATH", "x ATH", "ATH Value", "ATH Date",
        "min of 52", "max of 52",
        "min wk52", "max wk52",
        "min month12", "max month12",
        "bottom touch", "top touch",
        "bottom count", "top count",
        "number of min day touch", "number of max day touch",
    ]
    weekly  = []
    for wk in range(1, NUM_WEEKS + 1):
        weekly.extend([f"min wk{wk}", f"max wk{wk}", f"ratio wk{wk}"])
    monthly = []
    for mo in range(1, NUM_MONTHS + 1):
        monthly.extend([f"min month{mo}", f"max month{mo}", f"ration month{mo}"])
    daily1 = [f"day {d}" for d in range(1, NUM_DAYS + 1)]
    daily2 = [f"day {d}" for d in range(1, NUM_DAYS + 1)]
    tracking_dates = get_tracking_dates()
    return base + weekly + monthly + daily1 + daily2 + tracking_dates

# ============================================================
# HISTORICAL DATA READERS
# ============================================================

def get_previous_data_row(filename):
    """
    Reads Sheet 1 (Row Report) — assets are rows, headers are in a row
    where column A = 'ASSET'. Reads top-down and stops after the newest block.
    """
    counts        = {}
    tracking_data = {}
    if not os.path.exists(filename):
        return counts, tracking_data
    try:
        wb = load_workbook(filename, data_only=True)
        if SHEET_ROW not in wb.sheetnames:
            return counts, tracking_data
        ws = wb[SHEET_ROW]

        asset_col_idx    = None
        bot_cnt_col_idx  = None
        top_cnt_col_idx  = None
        tracking_col_indices = {}
        
        asset_header_count = 0

        for row in ws.iter_rows():
            if not row or not row[0].value: continue
            val0 = str(row[0].value).strip()
            
            if val0 == "ASSET":
                asset_header_count += 1
                if asset_header_count > 1:
                    break # Stop reading once we hit an older historical block
                
                for idx, cell in enumerate(row):
                    cval = str(cell.value).strip() if cell.value else ""
                    if   cval == "ASSET":        asset_col_idx   = idx
                    elif cval == "bottom count": bot_cnt_col_idx = idx
                    elif cval == "top count":    top_cnt_col_idx = idx
                    elif (cval.count("/") == 2 and
                          cval.replace("/", "").isdigit()):
                        tracking_col_indices[idx] = cval
                continue
                
            if asset_col_idx is not None:
                asset = row[asset_col_idx].value
                if asset:
                    b_val = row[bot_cnt_col_idx].value if bot_cnt_col_idx is not None else 0
                    t_val = row[top_cnt_col_idx].value if top_cnt_col_idx is not None else 0
                    counts[asset] = {
                        "bottom": int(b_val) if isinstance(b_val, (int, float)) else 0,
                        "top":    int(t_val) if isinstance(t_val, (int, float)) else 0,
                    }
                    t_data = {}
                    for col_idx, d_str in tracking_col_indices.items():
                        cell = row[col_idx]
                        if cell.value is not None and cell.value != "":
                            font_c = get_color_rgb(cell.font.color) if cell.font else None
                            fill_c = get_color_rgb(cell.fill.start_color) if cell.fill else None
                            t_data[d_str] = {
                                "value":      cell.value,
                                "font_color": font_c,
                                "fill_color": fill_c,
                                "is_bold":    cell.font.bold if cell.font else False,
                            }
                    tracking_data[asset] = t_data
    except Exception as e:
        print(f"  [Warning] Row sheet read failed: {e}")
    return counts, tracking_data


def get_previous_data_col(filename):
    """
    Reads Sheet 2 (Column Report) — transposed layout.
    Reads top-down and stops after the newest block.
    """
    counts        = {}
    tracking_data = {}
    if not os.path.exists(filename):
        return counts, tracking_data
    try:
        wb = load_workbook(filename, data_only=True)
        if SHEET_COL not in wb.sheetnames:
            return counts, tracking_data
        ws = wb[SHEET_COL]

        first_asset_row = None
        for r in range(1, ws.max_row + 1):
            val = str(ws.cell(row=r, column=3).value).strip()
            if val == "ASSET":
                first_asset_row = r
                break

        if first_asset_row is None:
            return counts, tracking_data

        metric_rows = {}
        for r in range(first_asset_row, ws.max_row + 1):
            m_val = str(ws.cell(row=r, column=3).value).strip()
            if m_val == "ASSET" and r > first_asset_row:
                break # Stop reading once we hit an older historical block
            if m_val:
                metric_rows[m_val] = r

        bot_cnt_row = metric_rows.get("bottom count")
        top_cnt_row = metric_rows.get("top count")

        tracking_row_indices = {}
        for m_val, r in metric_rows.items():
            if m_val.count("/") == 2 and m_val.replace("/", "").isdigit():
                tracking_row_indices[r] = m_val

        for c in range(4, ws.max_column + 1):
            asset = ws.cell(row=first_asset_row, column=c).value
            if asset:
                b_val = ws.cell(row=bot_cnt_row, column=c).value if bot_cnt_row else 0
                t_val = ws.cell(row=top_cnt_row, column=c).value if top_cnt_row else 0
                counts[asset] = {
                    "bottom": int(b_val) if isinstance(b_val, (int, float)) else 0,
                    "top":    int(t_val) if isinstance(t_val, (int, float)) else 0,
                }
                t_data = {}
                for r_idx, d_str in tracking_row_indices.items():
                    cell = ws.cell(row=r_idx, column=c)
                    if cell.value is not None and cell.value != "":
                        font_c = get_color_rgb(cell.font.color) if cell.font else None
                        fill_c = get_color_rgb(cell.fill.start_color) if cell.fill else None
                        t_data[d_str] = {
                            "value":      cell.value,
                            "font_color": font_c,
                            "fill_color": fill_c,
                            "is_bold":    cell.font.bold if cell.font else False,
                        }
                tracking_data[asset] = t_data
    except Exception as e:
        print(f"  [Warning] Column sheet read failed: {e}")
    return counts, tracking_data

# ============================================================
# BUCKET CALCULATORS
# ============================================================

def compute_buckets(prices, num_buckets, bucket_size):
    total   = len(prices)
    buckets = []
    for b in range(num_buckets):
        end_idx   = total - (b * bucket_size)
        start_idx = end_idx - bucket_size
        start_idx = max(start_idx, 0)
        end_idx   = max(end_idx,   0)
        if start_idx >= end_idx:
            buckets.append({"min": "", "max": "", "ratio": ""})
        else:
            chunk = prices[start_idx:end_idx]
            b_min = min(chunk)
            b_max = max(chunk)
            buckets.append({"min": b_min, "max": b_max,
                            "ratio": safe_ratio(b_min, b_max)})
    buckets.reverse()
    return buckets

def compute_daily_buckets(prices, num_days=30):
    window = prices[-num_days:]
    padded = [""] * (num_days - len(window)) + window
    return padded

def count_touch_days(daily_prices, touch_value, touch_type="min"):
    if touch_value == "" or touch_value is None:          return 0
    if not isinstance(touch_value, (int, float)):         return 0
    count = 0
    for price in daily_prices:
        if price == "" or not isinstance(price, (int, float)): continue
        if touch_type == "min":
            if price <= touch_value * 1.05: count += 1
        else:
            if price >= touch_value * 0.95: count += 1
    return count

# ============================================================
# MASTER DATA PROCESSOR
# ============================================================

def process_raw_data(name, prices, current_price,
                     previous_counts, previous_tracking_data,
                     is_stock=False):
    try:
        min_7d,  max_7d  = min(prices[-7:]),  max(prices[-7:])
        min_30d, max_30d = min(prices[-30:]), max(prices[-30:])
        min_1y,  max_1y  = min(prices),       max(prices)

        r_7d,  _,      _,     _      = calculate_metrics(min_7d,  max_7d)
        r_30d, _,      _,     _      = calculate_metrics(min_30d, max_30d)
        r_1y,  ref_1y, td_1y, tu_1y  = calculate_metrics(min_1y,  max_1y)

        ath_val  = ATH_DATA.get(name, {}).get("ath",  "N/A")
        ath_date = ATH_DATA.get(name, {}).get("date", "N/A")

        if (isinstance(current_price, (int, float)) and
                isinstance(ath_val, (int, float)) and ath_val > 0):
            pct_to_ath = current_price / ath_val
            x_ath      = ath_val / current_price
        else:
            pct_to_ath, x_ath = "ERROR", "ERROR"

        week_size  = 5  if is_stock else 7
        month_size = 21 if is_stock else 30

        weekly_buckets  = compute_buckets(prices, NUM_WEEKS,  week_size)
        monthly_buckets = compute_buckets(prices, NUM_MONTHS, month_size)
        daily_buckets   = compute_daily_buckets(prices, NUM_DAYS)

        valid_mins = [b['min'] for b in weekly_buckets
                      if isinstance(b['min'], (int, float))]
        valid_maxs = [b['max'] for b in weekly_buckets
                      if isinstance(b['max'], (int, float))]
        min_of_52  = min(valid_mins) if valid_mins else "N/A"
        max_of_52  = max(valid_maxs) if valid_maxs else "N/A"

        wk52    = weekly_buckets[-1]
        month12 = monthly_buckets[-1]

        min_wk52    = wk52.get("min",  "")
        max_wk52    = wk52.get("max",  "")
        min_month12 = month12.get("min", "")
        max_month12 = month12.get("max", "")

        if   min_wk52 == "" or min_month12 == "":  bottom_touch = "N/A"
        elif are_equal_prices(min_wk52, min_month12): bottom_touch = "Down"
        else:                                          bottom_touch = ""

        if   max_wk52 == "" or max_month12 == "":  top_touch = "N/A"
        elif are_equal_prices(max_wk52, max_month12): top_touch = "Up"
        else:                                          top_touch = ""

        prev_bot     = previous_counts.get(name, {}).get("bottom", 0)
        prev_top     = previous_counts.get(name, {}).get("top",    0)
        bottom_count = prev_bot + 1 if bottom_touch == "Down" else 0
        top_count    = prev_top + 1 if top_touch    == "Up"   else 0

        num_min_day_touch = count_touch_days(daily_buckets, min_month12, "min")
        num_max_day_touch = count_touch_days(daily_buckets, max_month12, "max")

        daily_week_flags = []
        for i, price in enumerate(daily_buckets):
            if price == "" or not isinstance(price, (int, float)):
                daily_week_flags.append((False, False)); continue
            days_from_end = NUM_DAYS - i - 1
            week_from_end = days_from_end // week_size
            wk_idx        = (NUM_WEEKS - 1) - week_from_end
            if wk_idx < 0:
                daily_week_flags.append((False, False)); continue
            wk_min    = weekly_buckets[wk_idx]['min']
            wk_max    = weekly_buckets[wk_idx]['max']
            is_wk_bot = (wk_min != "") and (price <= wk_min * 1.03)
            is_wk_top = (wk_max != "") and (price >= wk_max * 0.97)
            daily_week_flags.append((is_wk_bot, is_wk_top))

        asset_tracking = previous_tracking_data.get(name, {}).copy()
        now       = datetime.now()
        today_str = f"{now.month}/{now.day}/{now.year}"

        if bottom_touch == "Down" or top_touch == "Up":
            is_bottom       = (bottom_touch == "Down")
            highlight_color = None
            recorded_value  = ""
            if is_bottom:
                recorded_value = min_month12
                if (min_of_52 not in ["N/A", ""] and
                        min_wk52 != "" and min_month12 != ""):
                    if (are_equal_prices(min_of_52, min_wk52) and
                            are_equal_prices(min_wk52, min_month12)):
                        highlight_color = "FFFFFF00"
                        recorded_value  = min_of_52
            else:
                recorded_value = max_month12
                if (max_of_52 not in ["N/A", ""] and
                        max_wk52 != "" and max_month12 != ""):
                    if (are_equal_prices(max_of_52, max_wk52) and
                            are_equal_prices(max_wk52, max_month12)):
                        highlight_color = "FF00FF00"
                        recorded_value  = max_of_52
            asset_tracking[today_str] = {
                "value":      recorded_value,
                "font_color": "FFFF0000" if is_bottom else "FF0000FF",
                "fill_color": highlight_color,
                "is_bold":    True if highlight_color else False,
            }

        return {
            "name": name, "current": current_price, "is_stock": is_stock,
            "min_7d":  min_7d,  "max_7d":  max_7d,  "r_7d":  r_7d,
            "min_30d": min_30d, "max_30d": max_30d, "r_30d": r_30d,
            "min_1y":  min_1y,  "max_1y":  max_1y,
            "r_1y":  r_1y, "ref_1y": ref_1y, "td_1y": td_1y, "tu_1y": tu_1y,
            "pct_to_ath": pct_to_ath, "x_ath": x_ath,
            "ath": ath_val, "ath_date": ath_date,
            "min_of_52":   min_of_52,   "max_of_52":   max_of_52,
            "min_wk52":    min_wk52,    "max_wk52":    max_wk52,
            "min_month12": min_month12, "max_month12": max_month12,
            "bottom_touch":      bottom_touch,
            "top_touch":         top_touch,
            "bottom_count":      bottom_count,
            "top_count":         top_count,
            "num_min_day_touch": num_min_day_touch,
            "num_max_day_touch": num_max_day_touch,
            "weekly":           weekly_buckets,
            "monthly":          monthly_buckets,
            "daily":            daily_buckets,
            "daily_week_flags": daily_week_flags,
            "tracking_data":    asset_tracking,
        }

    except Exception as e:
        print(f"  Calculation error for {name}: {e}")
        return None

# ============================================================
# CONSOLE TABLE PRINTER
# ============================================================

def print_console_table(results, title):
    print(f"\n=== {title} ===")
    header = (
        f"{'ASSET':<8} | {'CURRENT':<12} | "
        f"{'7D MIN':<12} | {'7D MAX':<12} | "
        f"{'30D MIN':<12} | {'30D MAX':<12} | "
        f"{'1Y MIN':<12} | {'1Y MAX':<12} | "
        f"{'% ATH':<8} | {'x ATH':<7} | "
        f"{'MIN52':<12} | {'MAX52':<12} | "
        f"{'BOT':<5} | {'TOP':<5} | "
        f"{'#BOT':<5} | {'#TOP':<5}"
    )
    print(header)
    print("-" * len(header))
    for r in results:
        pct = (f"{r['pct_to_ath']:.1%}"
               if isinstance(r['pct_to_ath'], float) else r['pct_to_ath'])
        xat = (f"{r['x_ath']:.2f}x"
               if isinstance(r['x_ath'],      float) else r['x_ath'])
        print(
            f"{r['name']:<8} | {format_price(r['current']):<12} | "
            f"{format_price(r['min_7d']):<12} | {format_price(r['max_7d']):<12} | "
            f"{format_price(r['min_30d']):<12} | {format_price(r['max_30d']):<12} | "
            f"{format_price(r['min_1y']):<12} | {format_price(r['max_1y']):<12} | "
            f"{pct:<8} | {xat:<7} | "
            f"{format_price(r['min_of_52']):<12} | {format_price(r['max_of_52']):<12} | "
            f"{r['bottom_touch']:<5} | {r['top_touch']:<5} | "
            f"{r['num_min_day_touch']:<5} | {r['num_max_day_touch']:<5}"
        )

# ============================================================
# SHARED: build the flat data row for one asset result
# ============================================================

def build_asset_data_row(r, tracking_dates):
    row = [
        r['name'], format_excel_num(r['current']),
        format_excel_num(r['min_7d']),  format_excel_num(r['max_7d']),
        format_excel_num(r['r_7d']),
        format_excel_num(r['min_30d']), format_excel_num(r['max_30d']),
        format_excel_num(r['r_30d']),
        format_excel_num(r['min_1y']),  format_excel_num(r['max_1y']),
        format_excel_num(r['r_1y']),    format_excel_num(r['ref_1y']),
        format_excel_num(r['td_1y']),   format_excel_num(r['tu_1y']),
        format_excel_num(r['pct_to_ath']), format_excel_num(r['x_ath']),
        format_excel_num(r['ath']),     r['ath_date'],
        format_excel_num(r['min_of_52']),  format_excel_num(r['max_of_52']),
        format_excel_num(r['min_wk52']),   format_excel_num(r['max_wk52']),
        format_excel_num(r['min_month12']),format_excel_num(r['max_month12']),
        r['bottom_touch'], r['top_touch'],
        r['bottom_count'], r['top_count'],
        r['num_min_day_touch'], r['num_max_day_touch'],
    ]
    for b in r['weekly']:
        row.extend([format_excel_num(b['min']),
                    format_excel_num(b['max']),
                    format_excel_num(b['ratio'])])
    for b in r['monthly']:
        row.extend([format_excel_num(b['min']),
                    format_excel_num(b['max']),
                    format_excel_num(b['ratio'])])
    for price in r['daily']:
        row.append(format_excel_num(price))
    for price in r['daily']:
        row.append(format_excel_num(price))
    for d_str in tracking_dates:
        t_info = r['tracking_data'].get(d_str)
        row.append(format_excel_num(t_info["value"]) if t_info else "")
    return row

# ============================================================
# SHARED: apply all per-asset cell formatting
# ============================================================

def apply_asset_formatting(ws, current_row, r, headers,
                            tracking_dates,
                            idx_min_52, idx_max_52,
                            idx_bottom_touch, idx_top_touch,
                            idx_bottom_count, idx_top_count,
                            idx_min_wk52, idx_max_wk52,
                            idx_min_mo12, idx_max_mo12,
                            idx_day_start_1, idx_day_start_2,
                            idx_tracking_start,
                            idx_weekly_start,
                            center_cols,
                            red_font, blue_font,
                            bold_red_font, bold_blue_font,
                            thin_border, thick_right_border,
                            center_align,
                            col_offset=0):
    
    # Default thin border on every data cell
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=current_row, column=col_idx).border = thin_border

    # Week-grouping thick right borders in second daily group
    for i in range(NUM_DAYS):
        col_idx = idx_day_start_2 + i
        if r['is_stock']:
            if (i + 1) % 5 == 0:
                ws.cell(row=current_row, column=col_idx).border = thick_right_border
        else:
            if i % 7 == 1:
                ws.cell(row=current_row, column=col_idx).border = thick_right_border

    for col_idx in center_cols:
        ws.cell(row=current_row, column=col_idx).alignment = center_align

    if r['bottom_touch'] == "Down":
        fnt = bold_red_font if r['bottom_count'] > 1 else red_font
        ws.cell(row=current_row, column=idx_bottom_touch).font = fnt
        ws.cell(row=current_row, column=idx_min_wk52).font     = red_font
        ws.cell(row=current_row, column=idx_min_mo12).font     = red_font

    if r['top_touch'] == "Up":
        fnt = bold_blue_font if r['top_count'] > 1 else blue_font
        ws.cell(row=current_row, column=idx_top_touch).font = fnt
        ws.cell(row=current_row, column=idx_max_wk52).font  = blue_font
        ws.cell(row=current_row, column=idx_max_mo12).font  = blue_font

    if r['bottom_count'] == 1:
        ws.cell(row=current_row, column=idx_bottom_count).font = red_font
    elif r['bottom_count'] > 1:
        ws.cell(row=current_row, column=idx_bottom_count).font = bold_red_font

    if r['top_count'] == 1:
        ws.cell(row=current_row, column=idx_top_count).font = blue_font
    elif r['top_count'] > 1:
        ws.cell(row=current_row, column=idx_top_count).font = bold_blue_font

    if r['min_of_52'] not in ["N/A", ""]:
        if (are_equal_prices(r['min_of_52'], r['min_wk52']) or
                are_equal_prices(r['min_of_52'], r['min_month12'])):
            ws.cell(row=current_row, column=idx_min_52).font = bold_red_font

    if r['max_of_52'] not in ["N/A", ""]:
        if (are_equal_prices(r['max_of_52'], r['max_wk52']) or
                are_equal_prices(r['max_of_52'], r['max_month12'])):
            ws.cell(row=current_row, column=idx_max_52).font = bold_blue_font

    for i, price in enumerate(r['daily']):
        if price != "" and isinstance(price, (int, float)):
            is_bot = (r['min_month12'] != "" and price <= r['min_month12'] * 1.05)
            is_top = (r['max_month12'] != "" and price >= r['max_month12'] * 0.95)
            if is_bot:
                ws.cell(row=current_row,
                        column=idx_day_start_1 + i).font = red_font
            elif is_top:
                ws.cell(row=current_row,
                        column=idx_day_start_1 + i).font = blue_font

    for i, (is_wk_bot, is_wk_top) in enumerate(r['daily_week_flags']):
        if is_wk_bot:
            ws.cell(row=current_row,
                    column=idx_day_start_2 + i).font = red_font
        elif is_wk_top:
            ws.cell(row=current_row,
                    column=idx_day_start_2 + i).font = blue_font

    # --- NEW: Propagate daily cell format (red/blue) to weekly/monthly/tracking cells ---
    daily_match_formats = []
    for i, price in enumerate(r['daily']):
        if price != "" and isinstance(price, (int, float)):
            is_bot = (r['min_month12'] != "" and price <= r['min_month12'] * 1.05)
            is_top = (r['max_month12'] != "" and price >= r['max_month12'] * 0.95)
            if is_bot:
                daily_match_formats.append((price, red_font))
            elif is_top:
                daily_match_formats.append((price, blue_font))

    end_col = idx_tracking_start + len(tracking_dates) if idx_tracking_start else idx_day_start_2 + NUM_DAYS
    if idx_weekly_start and daily_match_formats:
        for col_idx in range(idx_weekly_start, end_col):
            cell = ws.cell(row=current_row, column=col_idx)
            val = cell.value
            parsed_val = parse_price_for_match(val)
            if parsed_val is not None:
                for d_val, d_font in daily_match_formats:
                    if are_equal_prices(parsed_val, d_val):
                        cell.font = d_font
                        break

    # --- NEW: Match tracking formats to weekly/monthly/daily cells ---
    tracking_formats = []
    if idx_tracking_start:
        for i, d_str in enumerate(tracking_dates):
            col_idx = idx_tracking_start + i
            t_info  = r['tracking_data'].get(d_str)
            if t_info:
                # Apply format to the tracking date cell itself
                cell    = ws.cell(row=current_row, column=col_idx)
                f_color = t_info.get("font_color")
                is_bold = t_info.get("is_bold", False)
                if f_color:
                    cell.font = Font(color=f_color, bold=is_bold)
                elif is_bold:
                    cell.font = Font(bold=True)
                bg_color = t_info.get("fill_color")
                if bg_color and bg_color != "00000000":
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                
                # Store format info for matching in earlier columns
                t_val = t_info.get("value")
                if t_val != "" and t_val is not None:
                    if f_color or bg_color or is_bold:
                        tracking_formats.append((t_val, t_info))

    # Apply matched formats to weekly/monthly/daily blocks
    if tracking_formats and idx_weekly_start and idx_tracking_start:
        for col_idx in range(idx_weekly_start, end_col):
            cell = ws.cell(row=current_row, column=col_idx)
            val = cell.value
            
            parsed_val = parse_price_for_match(val)
            if parsed_val is not None:
                for t_val, t_info in tracking_formats:
                    parsed_t_val = parse_price_for_match(t_val)
                    if parsed_t_val is not None and are_equal_prices(parsed_val, parsed_t_val):
                        # Match found! Apply tracking format to this cell
                        f_color = t_info.get("font_color")
                        is_bold = t_info.get("is_bold", False)
                        if f_color:
                            cell.font = Font(color=f_color, bold=is_bold)
                        elif is_bold:
                            cell.font = Font(bold=True)
                            
                        bg_color = t_info.get("fill_color")
                        if bg_color and bg_color != "00000000":
                            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                        break # First match wins

# ============================================================
# SHEET 1 WRITER — ROW REPORT
# ============================================================

def write_sheet_row(ws, results):
    """Writes the Row Report sheet (assets as rows) putting newest data at the top."""
    headers        = build_excel_headers()
    tracking_dates = get_tracking_dates()

    now      = datetime.now()
    time_str = now.strftime("%H:%M:%S")
    date_str = now.strftime("%m/%d/%Y")

    base_len    = 30
    weekly_len  = NUM_WEEKS  * 3
    monthly_len = NUM_MONTHS * 3
    idx_daily2_start       = base_len + weekly_len + monthly_len + NUM_DAYS
    offset_to_second_daily = idx_daily2_start

    # Row 1 — timestamp + 7-day cycle labels
    row1_days = [f"day {((1 - i) % 7) + 1}" for i in range(NUM_DAYS)]
    row1 = ([time_str, date_str]
            + [""] * (offset_to_second_daily - 2)
            + row1_days
            + [""] * len(tracking_dates))

    # Row 2 — 5-day cycle labels
    row2_days = [f"day {(i % 5) + 1}" for i in range(NUM_DAYS)]
    row2 = ([""] * offset_to_second_daily
            + row2_days
            + [""] * len(tracking_dates))

    # Build the full block of data in memory
    block = [row1, row2, headers]
    for r in results:
        block.append(build_asset_data_row(r, tracking_dates))

    # If the sheet already contains data, push it down to make room at the top
    has_existing_data = ws.max_row > 1 or (ws.max_row == 1 and ws.cell(1, 1).value is not None)
    if has_existing_data:
        # Insert enough rows for the new block + 1 blank separator row
        ws.insert_rows(1, len(block) + 1)

    # Write the new block into the top rows
    for r_idx, row_data in enumerate(block, start=1):
        for c_idx, val in enumerate(row_data, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    row1_idx   = 1
    row2_idx   = 2
    header_row = 3

    # ── Styles ──────────────────────────────────────────────
    center_align       = Alignment(horizontal="center")
    # FIX: Strict 8-character aRGB hex colors required by Excel OOXML standard
    red_font           = Font(color="FFFF0000")
    blue_font          = Font(color="FF0000FF")
    bold_red_font      = Font(color="FFFF0000", bold=True)
    bold_blue_font     = Font(color="FF0000FF", bold=True)
    
    thin_border        = Border(left=Side(style='thin'),   right=Side(style='thin'),
                                top=Side(style='thin'),    bottom=Side(style='thin'))
    thick_right_border = Border(left=Side(style='thin'),   right=Side(style='medium'),
                                top=Side(style='thin'),    bottom=Side(style='thin'))

    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=col_idx).border = thin_border

    idx_min_52        = headers.index("min of 52")        + 1
    idx_max_52        = headers.index("max of 52")        + 1
    idx_bottom_touch  = headers.index("bottom touch")     + 1
    idx_top_touch     = headers.index("top touch")        + 1
    idx_bottom_count  = headers.index("bottom count")     + 1
    idx_top_count     = headers.index("top count")        + 1
    idx_min_wk52      = headers.index("min wk52")         + 1
    idx_max_wk52      = headers.index("max wk52")         + 1
    idx_min_mo12      = headers.index("min month12")      + 1
    idx_max_mo12      = headers.index("max month12")      + 1
    idx_day_start_1   = headers.index("day 1")            + 1
    idx_day_start_2   = idx_day_start_1 + NUM_DAYS
    idx_tracking_start= (headers.index(tracking_dates[0]) + 1
                         if tracking_dates else None)
    idx_weekly_start  = headers.index("min wk1") + 1

    center_cols = [
        idx_bottom_touch, idx_top_touch,
        idx_bottom_count, idx_top_count,
        headers.index("number of min day touch") + 1,
        headers.index("number of max day touch") + 1,
    ]

    # Thick right borders on header rows for second daily group
    for i in range(NUM_DAYS):
        col_idx = idx_day_start_2 + i
        if i % 7 == 1:
            ws.cell(row=row1_idx, column=col_idx).border = thick_right_border
        if (i + 1) % 5 == 0:
            ws.cell(row=row2_idx, column=col_idx).border = thick_right_border

    # ── Data rows formatting ─────────────────────────────────
    for i, r in enumerate(results):
        current_row = 4 + i  # Data starts at row 4
        apply_asset_formatting(
            ws, current_row, r, headers, tracking_dates,
            idx_min_52, idx_max_52,
            idx_bottom_touch, idx_top_touch,
            idx_bottom_count, idx_top_count,
            idx_min_wk52, idx_max_wk52,
            idx_min_mo12, idx_max_mo12,
            idx_day_start_1, idx_day_start_2,
            idx_tracking_start,
            idx_weekly_start,
            center_cols,
            red_font, blue_font, bold_red_font, bold_blue_font,
            thin_border, thick_right_border, center_align,
        )

    # Freeze panes so headers (Rows 1-3) and asset names (Cols A-B) remain visible when scrolling down
    ws.freeze_panes = 'C4'

# ============================================================
# SHEET 2 WRITER — COLUMN REPORT (TRANSPOSED)
# ============================================================

def write_sheet_col(ws, results):
    """Writes the Column Report sheet (assets as columns) putting newest data at the top."""
    headers        = build_excel_headers()
    tracking_dates = get_tracking_dates()

    now      = datetime.now()
    time_str = now.strftime("%H:%M:%S")
    date_str = now.strftime("%m/%d/%Y")

    base_len    = 30
    weekly_len  = NUM_WEEKS  * 3
    monthly_len = NUM_MONTHS * 3
    idx_daily2_start       = base_len + weekly_len + monthly_len + NUM_DAYS
    offset_to_second_daily = idx_daily2_start

    # Build pre-transposed rows
    row1_days = [f"day {((1 - i) % 7) + 1}" for i in range(NUM_DAYS)]
    row1 = ([time_str, date_str]
            + [""] * (offset_to_second_daily - 2)
            + row1_days
            + [""] * len(tracking_dates))

    row2_days = [f"day {(i % 5) + 1}" for i in range(NUM_DAYS)]
    row2 = ([""] * offset_to_second_daily
            + row2_days
            + [""] * len(tracking_dates))

    row3      = headers
    data_rows = [build_asset_data_row(r, tracking_dates) for r in results]

    block            = [row1, row2, row3] + data_rows
    transposed_block = list(map(list, zip(*block)))

    # If the sheet already contains data, push it down to make room at the top
    has_existing_data = ws.max_row > 1 or (ws.max_row == 1 and ws.cell(1, 1).value is not None)
    if has_existing_data:
        # Insert enough rows for the new block + 1 blank separator row
        ws.insert_rows(1, len(transposed_block) + 1)

    start_row = 1

    # Write the block at the top
    for r_idx, t_row in enumerate(transposed_block, start=start_row):
        for c_idx, val in enumerate(t_row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # ── Styles ──────────────────────────────────────────────
    center_align        = Alignment(horizontal="center")
    # FIX: Strict 8-character aRGB hex colors required by Excel OOXML standard
    red_font            = Font(color="FFFF0000")
    blue_font           = Font(color="FF0000FF")
    bold_red_font       = Font(color="FFFF0000", bold=True)
    bold_blue_font      = Font(color="FF0000FF", bold=True)
    
    thin_border         = Border(left=Side(style='thin'),  right=Side(style='thin'),
                                 top=Side(style='thin'),   bottom=Side(style='thin'))
    thick_bottom_border = Border(left=Side(style='thin'),  right=Side(style='thin'),
                                 top=Side(style='thin'),   bottom=Side(style='medium'))

    def get_t_row(orig_col_idx):
        return start_row + orig_col_idx - 1

    # Thin border on every cell in the block
    for r_idx in range(start_row, start_row + len(transposed_block)):
        for c_idx in range(1, len(transposed_block[0]) + 1):
            ws.cell(row=r_idx, column=c_idx).border = thin_border

    idx_min_52        = headers.index("min of 52")        + 1
    idx_max_52        = headers.index("max of 52")        + 1
    idx_bottom_touch  = headers.index("bottom touch")     + 1
    idx_top_touch     = headers.index("top touch")        + 1
    idx_bottom_count  = headers.index("bottom count")     + 1
    idx_top_count     = headers.index("top count")        + 1
    idx_min_wk52      = headers.index("min wk52")         + 1
    idx_max_wk52      = headers.index("max wk52")         + 1
    idx_min_mo12      = headers.index("min month12")      + 1
    idx_max_mo12      = headers.index("max month12")      + 1
    idx_day_start_1   = headers.index("day 1")            + 1
    idx_day_start_2   = idx_day_start_1 + NUM_DAYS
    idx_tracking_start= (headers.index(tracking_dates[0]) + 1
                         if tracking_dates else None)
    idx_weekly_start  = headers.index("min wk1") + 1

    center_orig_cols = [
        idx_bottom_touch, idx_top_touch,
        idx_bottom_count, idx_top_count,
        headers.index("number of min day touch") + 1,
        headers.index("number of max day touch") + 1,
    ]

    # Thick borders on header cols (cols 1 & 2) for second daily group
    for i in range(NUM_DAYS):
        orig_col  = idx_day_start_2 + i
        t_row_idx = get_t_row(orig_col)
        if i % 7 == 1:
            ws.cell(row=t_row_idx, column=1).border = thick_bottom_border
        if (i + 1) % 5 == 0:
            ws.cell(row=t_row_idx, column=2).border = thick_bottom_border

    # Per-asset column formatting
    for asset_idx, r in enumerate(results):
        t_col = 4 + asset_idx   # cols 1-3 = meta; col 4 = first asset

        # Week-grouping borders in second daily group
        for i in range(NUM_DAYS):
            orig_col  = idx_day_start_2 + i
            t_row_idx = get_t_row(orig_col)
            if r['is_stock']:
                if (i + 1) % 5 == 0:
                    ws.cell(row=t_row_idx, column=t_col).border = thick_bottom_border
            else:
                if i % 7 == 1:
                    ws.cell(row=t_row_idx, column=t_col).border = thick_bottom_border

        for orig_col in center_orig_cols:
            ws.cell(row=get_t_row(orig_col),
                    column=t_col).alignment = center_align

        if r['bottom_touch'] == "Down":
            fnt = bold_red_font if r['bottom_count'] > 1 else red_font
            ws.cell(row=get_t_row(idx_bottom_touch), column=t_col).font = fnt
            ws.cell(row=get_t_row(idx_min_wk52),     column=t_col).font = red_font
            ws.cell(row=get_t_row(idx_min_mo12),     column=t_col).font = red_font

        if r['top_touch'] == "Up":
            fnt = bold_blue_font if r['top_count'] > 1 else blue_font
            ws.cell(row=get_t_row(idx_top_touch), column=t_col).font = fnt
            ws.cell(row=get_t_row(idx_max_wk52),  column=t_col).font = blue_font
            ws.cell(row=get_t_row(idx_max_mo12),  column=t_col).font = blue_font

        if r['bottom_count'] == 1:
            ws.cell(row=get_t_row(idx_bottom_count), column=t_col).font = red_font
        elif r['bottom_count'] > 1:
            ws.cell(row=get_t_row(idx_bottom_count), column=t_col).font = bold_red_font

        if r['top_count'] == 1:
            ws.cell(row=get_t_row(idx_top_count), column=t_col).font = blue_font
        elif r['top_count'] > 1:
            ws.cell(row=get_t_row(idx_top_count), column=t_col).font = bold_blue_font

        if r['min_of_52'] not in ["N/A", ""]:
            if (are_equal_prices(r['min_of_52'], r['min_wk52']) or
                    are_equal_prices(r['min_of_52'], r['min_month12'])):
                ws.cell(row=get_t_row(idx_min_52), column=t_col).font = bold_red_font

        if r['max_of_52'] not in ["N/A", ""]:
            if (are_equal_prices(r['max_of_52'], r['max_wk52']) or
                    are_equal_prices(r['max_of_52'], r['max_month12'])):
                ws.cell(row=get_t_row(idx_max_52), column=t_col).font = bold_blue_font

        for i, price in enumerate(r['daily']):
            if price != "" and isinstance(price, (int, float)):
                is_bot = (r['min_month12'] != "" and price <= r['min_month12'] * 1.05)
                is_top = (r['max_month12'] != "" and price >= r['max_month12'] * 0.95)
                if is_bot:
                    ws.cell(row=get_t_row(idx_day_start_1 + i),
                            column=t_col).font = red_font
                elif is_top:
                    ws.cell(row=get_t_row(idx_day_start_1 + i),
                            column=t_col).font = blue_font

        for i, (is_wk_bot, is_wk_top) in enumerate(r['daily_week_flags']):
            if is_wk_bot:
                ws.cell(row=get_t_row(idx_day_start_2 + i),
                        column=t_col).font = red_font
            elif is_wk_top:
                ws.cell(row=get_t_row(idx_day_start_2 + i),
                        column=t_col).font = blue_font

        # --- NEW: Propagate daily cell format (red/blue) to weekly/monthly/tracking cells ---
        daily_match_formats = []
        for i, price in enumerate(r['daily']):
            if price != "" and isinstance(price, (int, float)):
                is_bot = (r['min_month12'] != "" and price <= r['min_month12'] * 1.05)
                is_top = (r['max_month12'] != "" and price >= r['max_month12'] * 0.95)
                if is_bot:
                    daily_match_formats.append((price, red_font))
                elif is_top:
                    daily_match_formats.append((price, blue_font))

        end_orig_col = idx_tracking_start + len(tracking_dates) if idx_tracking_start else idx_day_start_2 + NUM_DAYS
        if idx_weekly_start and daily_match_formats:
            for orig_col in range(idx_weekly_start, end_orig_col):
                t_row_idx = get_t_row(orig_col)
                cell = ws.cell(row=t_row_idx, column=t_col)
                val = cell.value
                parsed_val = parse_price_for_match(val)
                if parsed_val is not None:
                    for d_val, d_font in daily_match_formats:
                        if are_equal_prices(parsed_val, d_val):
                            cell.font = d_font
                            break

        # --- NEW: Match tracking formats to weekly/monthly/daily cells ---
        tracking_formats = []
        if idx_tracking_start:
            for i, d_str in enumerate(tracking_dates):
                orig_col = idx_tracking_start + i
                t_info   = r['tracking_data'].get(d_str)
                if t_info:
                    # Apply format to the tracking date cell itself
                    cell    = ws.cell(row=get_t_row(orig_col), column=t_col)
                    f_color = t_info.get("font_color")
                    is_bold = t_info.get("is_bold", False)
                    if f_color:
                        cell.font = Font(color=f_color, bold=is_bold)
                    elif is_bold:
                        cell.font = Font(bold=True)
                    bg_color = t_info.get("fill_color")
                    if bg_color and bg_color != "00000000":
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    
                    # Store format info for matching in earlier columns
                    t_val = t_info.get("value")
                    if t_val != "" and t_val is not None:
                        if f_color or bg_color or is_bold:
                            tracking_formats.append((t_val, t_info))

        # Apply matched formats to weekly/monthly/daily blocks
        if tracking_formats and idx_weekly_start and idx_tracking_start:
            for orig_col in range(idx_weekly_start, end_orig_col):
                t_row_idx = get_t_row(orig_col)
                cell = ws.cell(row=t_row_idx, column=t_col)
                val = cell.value
                
                parsed_val = parse_price_for_match(val)
                if parsed_val is not None:
                    for t_val, t_info in tracking_formats:
                        parsed_t_val = parse_price_for_match(t_val)
                        if parsed_t_val is not None and are_equal_prices(parsed_val, parsed_t_val):
                            # Match found! Apply tracking format to this cell
                            f_color = t_info.get("font_color")
                            is_bold = t_info.get("is_bold", False)
                            if f_color:
                                cell.font = Font(color=f_color, bold=is_bold)
                            elif is_bold:
                                cell.font = Font(bold=True)
                                
                            bg_color = t_info.get("fill_color")
                            if bg_color and bg_color != "00000000":
                                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                            break # First match wins

    # Freeze: lock rows 1-2 (ASSET + CURRENT) and cols A-C (meta cols)
    ws.freeze_panes = 'D3'

# ============================================================
# MASTER EXCEL WRITER  (opens / creates the workbook once)
# ============================================================

def write_excel(results, filename):
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        # Remove the default blank sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ── Ensure both sheets exist ─────────────────────────────
    if SHEET_ROW not in wb.sheetnames:
        wb.create_sheet(SHEET_ROW, 0)
    if SHEET_COL not in wb.sheetnames:
        wb.create_sheet(SHEET_COL, 1)

    ws_row = wb[SHEET_ROW]
    ws_col = wb[SHEET_COL]

    print(f"  Writing Sheet 1: '{SHEET_ROW}' ...")
    write_sheet_row(ws_row, results)

    print(f"  Writing Sheet 2: '{SHEET_COL}' ...")
    write_sheet_col(ws_col, results)

    try:
        wb.save(filename)
        print(f"  ✔ Saved: '{filename}'")
    except Exception as e:
        print(f"  ✘ Failed to write '{filename}': {e}")
        print("    (Make sure the Excel file is not currently open!)")

# ============================================================
# PHASE 1 — CRYPTO FETCHER  (CoinGecko, free tier)
# ============================================================

def fetch_crypto_data(previous_counts, previous_tracking_data):
    print("\n" + "=" * 65)
    print("  PHASE 1: Fetching CRYPTOCURRENCY data (CoinGecko)")
    print("=" * 65)

    req_headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        )
    }
    results = []

    for coin in COIN_LIST:
        coin_upper = coin.upper()
        url    = f"https://api.coingecko.com/api/v3/coins/{coin}/market_chart"
        params = {"vs_currency": "usd", "days": "365"}

        for attempt in range(3):
            try:
                print(f"  Fetching {coin_upper:<12} ... ", end="", flush=True)
                resp = requests.get(url, params=params,
                                    headers=req_headers, timeout=15)
                resp.raise_for_status()
                data = resp.json()

                if 'prices' not in data or not data['prices']:
                    print("FAILED (no price data)")
                    break

                prices = [p[1] for p in data['prices']]
                res    = process_raw_data(
                    coin_upper, prices, prices[-1],
                    previous_counts, previous_tracking_data, is_stock=False)
                if res:
                    results.append(res)
                    print("OK")
                time.sleep(5)
                break

            except requests.exceptions.HTTPError:
                if resp.status_code == 429:
                    wait = 20 * (attempt + 1)
                    print(f"Rate limited — waiting {wait}s ...")
                    time.sleep(wait)
                else:
                    print(f"HTTP Error {resp.status_code}")
                    break
            except Exception as e:
                print(f"ERROR: {e}")
                break

    return results

# ============================================================
# PHASE 2 — STOCK FETCHER  (yfinance)
# ============================================================

def fetch_stock_data(previous_counts, previous_tracking_data):
    print("\n" + "=" * 65)
    print("  PHASE 2: Fetching STOCK data (yfinance / Yahoo Finance)")
    print(f"  Stocks to fetch: {len(STOCK_LIST)}")
    print("=" * 65)

    results = []
    skipped = []

    for stock in STOCK_LIST:
        stock_upper = stock.upper()
        try:
            print(f"  Fetching {stock_upper:<6} ... ", end="", flush=True)
            ticker = yf.Ticker(stock_upper)
            hist   = ticker.history(period="1y")

            if hist.empty:
                print("SKIPPED (no data — may be delisted or invalid ticker)")
                skipped.append(stock_upper)
                continue

            prices        = hist['Close'].tolist()
            current_price = ticker.fast_info['last_price']

            res = process_raw_data(
                stock_upper, prices, current_price,
                previous_counts, previous_tracking_data, is_stock=True)
            if res:
                results.append(res)
                print(f"OK  ({len(prices)} days, current=${current_price:.2f})")
            time.sleep(1)

        except Exception as e:
            print(f"ERROR: {e}")
            skipped.append(stock_upper)

    if skipped:
        print(f"\n  ⚠ Skipped: {', '.join(skipped)}")

    return results

# ============================================================
# MAIN ENTRY POINT
# ============================================================

def fetch_all_data():
    # Read previous data from both sheets of the combined report
    prev_counts_row, prev_tracking_row = get_previous_data_row(EXCEL_FILE)
    prev_counts_col, prev_tracking_col = get_previous_data_col(EXCEL_FILE)

    # Merge: prefer row-sheet data; fall back to col-sheet data
    previous_counts   = {**prev_counts_col,   **prev_counts_row}
    previous_tracking = {**prev_tracking_col,  **prev_tracking_row}

    crypto_results = fetch_crypto_data(previous_counts, previous_tracking)
    stock_results  = fetch_stock_data(previous_counts,  previous_tracking)
    all_results    = crypto_results + stock_results

    print_console_table(crypto_results, "CRYPTO — Summary")
    print_console_table(stock_results,  "STOCKS — Summary")

    print("\n" + "=" * 65)
    print("  Writing Excel output ...")
    print("=" * 65)

    write_excel(all_results, EXCEL_FILE)

    total_metrics = len(build_excel_headers())
    print("\n✔ All done!")
    print(f"  Output file    : {EXCEL_FILE}")
    print(f"  Sheet 1        : '{SHEET_ROW}'  (assets = rows)")
    print(f"  Sheet 2        : '{SHEET_COL}'  (assets = columns)")
    print(f"  Crypto total   : {len(crypto_results)} / {len(COIN_LIST)}")
    print(f"  Stocks total   : {len(stock_results)} / {len(STOCK_LIST)}")
    print(f"  Assets total   : {len(all_results)}")
    print(f"  Metrics/asset  : {total_metrics}")

if __name__ == "__main__":
    fetch_all_data()